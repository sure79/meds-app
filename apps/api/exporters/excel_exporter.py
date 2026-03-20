"""Excel exporter for load balance calculations using openpyxl."""
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from ..engines.load_balance_engine import calculate_load_balance, CONDITION_LABELS


def export_load_balance_excel(req) -> bytes:
    """
    Export load balance calculation results to Excel.

    Layout:
    - Header section: project info
    - Matrix table: loads as rows, conditions as columns
    - Summary rows at the bottom
    - Conditional formatting for load percentages
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Load Balance"

    # Styles
    header_font = Font(name="Arial", size=14, bold=True)
    sub_font = Font(name="Arial", size=10)
    col_header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    col_header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    data_font = Font(name="Arial", size=10)
    summary_font = Font(name="Arial", size=10, bold=True)
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    amber_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Header section
    ws.merge_cells("A1:H1")
    ws["A1"] = "ELECTRICAL LOAD BALANCE"
    ws["A1"].font = header_font

    ws["A2"] = "Project:"
    ws["B2"] = req.project_name
    ws["A3"] = "Vessel:"
    ws["B3"] = req.vessel_name
    ws["A4"] = "Class Society:"
    ws["B4"] = req.class_society
    ws["A5"] = "System Voltage:"
    ws["B5"] = f"{req.system_voltage}V / {req.frequency}Hz"
    ws["A6"] = "Date:"
    ws["B6"] = datetime.now().strftime("%Y-%m-%d")

    for r in range(2, 7):
        ws[f"A{r}"].font = Font(name="Arial", size=10, bold=True)
        ws[f"B{r}"].font = sub_font

    # Determine conditions
    conditions = []
    for load in req.loads:
        for cond in load.load_factors:
            if cond not in conditions:
                conditions.append(cond)
    if not conditions:
        conditions = ["sea_going", "maneuvering", "port_loading", "port_idle", "emergency"]

    # Calculate results
    results = calculate_load_balance(
        req.generators, req.loads, req.buses, conditions,
        margin=None, class_society=req.class_society
    )

    # Table header row
    row = 8
    headers = ["No.", "Equipment", "Type", "Rated kW", "PF", "Bus"]
    for i, cond in enumerate(conditions):
        label = CONDITION_LABELS.get(cond, cond)
        headers.append(label)

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = col_header_font
        cell.fill = col_header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border

    # Data rows
    row += 1
    for idx, load in enumerate(req.loads, 1):
        ws.cell(row=row, column=1, value=idx).font = data_font
        ws.cell(row=row, column=2, value=load.name).font = data_font
        ws.cell(row=row, column=3, value=load.type).font = data_font
        ws.cell(row=row, column=4, value=load.rated_power_kw).font = data_font
        ws.cell(row=row, column=5, value=load.power_factor).font = data_font
        ws.cell(row=row, column=6, value=load.bus_id).font = data_font

        for ci, cond in enumerate(conditions):
            lf = load.load_factors.get(cond, 0.0)
            df = load.diversity_factor
            value = round(load.rated_power_kw * lf * df, 2)
            cell = ws.cell(row=row, column=7 + ci, value=value)
            cell.font = data_font
            cell.number_format = "0.00"
            cell.alignment = Alignment(horizontal="center")

        for col_idx in range(1, len(headers) + 1):
            ws.cell(row=row, column=col_idx).border = thin_border

        row += 1

    # Summary rows
    summary_start = row
    summary_labels = [
        "Total Load (kW)",
        "Total Load (kVA)",
        "Generator Capacity (kW)",
        "Load (%)",
        "Margin (kW)",
        "Margin (%)",
        "Status",
    ]

    result_map = {r["condition"]: r for r in results}

    for si, label in enumerate(summary_labels):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        cell = ws.cell(row=row, column=1, value=label)
        cell.font = summary_font
        cell.alignment = Alignment(horizontal="right")
        cell.border = thin_border
        for merge_col in range(2, 7):
            ws.cell(row=row, column=merge_col).border = thin_border

        for ci, cond in enumerate(conditions):
            r_data = result_map.get(cond, {})
            if si == 0:
                val = r_data.get("total_load_kw", 0)
            elif si == 1:
                val = r_data.get("total_load_kva", 0)
            elif si == 2:
                val = r_data.get("total_gen_capacity_kw", 0)
            elif si == 3:
                val = r_data.get("load_percentage", 0)
            elif si == 4:
                val = r_data.get("margin_kw", 0)
            elif si == 5:
                val = r_data.get("margin_percentage", 0)
            elif si == 6:
                val = "OK" if r_data.get("is_acceptable", False) else "NG"
            else:
                val = ""

            cell = ws.cell(row=row, column=7 + ci, value=val)
            cell.font = summary_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

            # Conditional formatting for load percentage and status
            if si == 3:  # Load %
                if isinstance(val, (int, float)):
                    if val < 80:
                        cell.fill = green_fill
                    elif val <= 90:
                        cell.fill = amber_fill
                    else:
                        cell.fill = red_fill
                cell.number_format = "0.00"
            elif si == 6:  # Status
                if val == "OK":
                    cell.fill = green_fill
                else:
                    cell.fill = red_fill

        row += 1

    # Set column widths
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 8
    ws.column_dimensions["F"].width = 12
    for ci in range(len(conditions)):
        col_letter = get_column_letter(7 + ci)
        ws.column_dimensions[col_letter].width = 18

    # Write to bytes
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()
